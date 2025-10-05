from openpyxl import load_workbook
from utils.DateTimeUtils import DateTimeUtils

datetime_utils = DateTimeUtils()

class ExcelUtils:

    def load_data(self, data_path):
        workbook = load_workbook(data_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        rows = []

        for row in sheet.iter_rows(min_row=2):
            clean_row = [datetime_utils.clean_value(cell.value) for cell in row]
            rows.append(clean_row)
        return headers, rows