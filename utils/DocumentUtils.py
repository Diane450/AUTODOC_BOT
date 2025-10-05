from openpyxl import load_workbook
from docxtpl import DocxTemplate
from docxcompose.composer import Composer
from docx import Document
from aiogram.types import Document as Doc
import pandas as pd
import datetime
import zipfile
import os
import shutil
from aiogram import Bot

class DocumentUtils:

    def _load_data(self, data_path: str):
        """Загружает данные из Excel или Google Sheets и корректно обрабатывает даты/время"""
        if "docs.google.com/spreadsheets" in data_path:
            sheet_id = data_path.split("/d/")[1].split("/")[0]
            csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
            df = pd.read_csv(csv_url)
            headers = list(df.columns)
            rows = df.values.tolist()

            # Преобразуем даты/время в строки
            clean_rows = []
            for row in rows:
                clean_row = []
                for value in row:
                    if isinstance(value, str):
                        clean_row.append(value)
                    elif pd.isna(value):
                        clean_row.append("")
                    elif isinstance(value, (datetime.date, datetime.datetime)):
                        # Если есть время
                        if isinstance(value, datetime.datetime):
                            clean_row.append(value.strftime("%d.%m.%Y %H:%M"))
                        else:
                            clean_row.append(value.strftime("%d.%m.%Y"))
                    else:
                        clean_row.append(value)
                clean_rows.append(clean_row)
            rows = clean_rows

        else:
            workbook = load_workbook(data_path, data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            rows = []
            for row in sheet.iter_rows(min_row=2):
                clean_row = []
                for cell in row:
                    value = cell.value
                    if isinstance(value, datetime.datetime):
                        fmt = str(cell.number_format).lower()
                        if "h" in fmt or "ч" in fmt:
                            value = value.strftime("%d.%m.%Y %H:%M")
                        else:
                            value = value.strftime("%d.%m.%Y")
                    elif isinstance(value, datetime.date):
                        value = value.strftime("%d.%m.%Y")
                    clean_row.append(value)
                rows.append(clean_row)

        return headers, rows

    def generate_document(self, template_path: str, data_path: str, output_type: str):
        os.makedirs("temp", exist_ok=True)

        headers, rows = self._load_data(data_path)
        temp_files = []

        for row in rows:
            context = dict(zip(headers, row))
            template = DocxTemplate(template_path)
            template.render(context)

            timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename_key = context.get("DOC_NAME", f"{timestamp}")
            output_filename = f"temp/{filename_key}.docx"

            template.save(output_filename)
            temp_files.append(output_filename)
            print(f"✅ Создан файл: {output_filename}")

        if output_type == "zip":
            zip_filename = f"temp/documents_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.zip"
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file in temp_files:
                    zipf.write(file, os.path.basename(file))
            return zip_filename

        elif output_type == "single":
            master_doc = Document(temp_files[0])
            composer = Composer(master_doc)
            for f in temp_files[1:]:
                composer.append(Document(f))
            combined_filename = f"temp/combined_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.docx"
            composer.save(combined_filename)
            return combined_filename


    def create_temp_path(self, doc: Doc):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_path = f"temp/{timestamp} - {doc.file_name}"
        return file_path

    async def download_file(self, bot: Bot, doc: Doc):
        file = await bot.get_file(doc.file_id)
        temp_file_path = self.create_temp_path(doc)
        await bot.download_file(file.file_path, temp_file_path)
        return temp_file_path
